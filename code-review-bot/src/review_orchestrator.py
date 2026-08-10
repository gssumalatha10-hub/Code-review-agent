#!/usr/bin/env python3
"""
Code Review Orchestrator for Eclipse iceoryx
Manages build, analysis, result collection, and reporting
"""

import os
import json
import subprocess
import yaml
import sys
import argparse
import shutil
from pathlib import Path
from typing import List, Dict, Any, Set, Tuple
from dataclasses import dataclass, asdict
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    get_column_letter = None

# SARIF result container
@dataclass
class Finding:
    rule_id: str
    tool: str
    severity: str
    confidence: str
    file: str
    line: int
    column: int
    message: str
    standard: str
    is_new: bool
    fingerprint: str
    
    def to_sarif_result(self) -> Dict[str, Any]:
        """Convert to SARIF result object"""
        return {
            "ruleId": self.rule_id,
            "message": {"text": self.message},
            "level": self.severity.lower(),
            "locations": [{
                "physicalLocation": {
                    "artifactLocation": {"uri": self.file},
                    "region": {
                        "startLine": self.line,
                        "startColumn": self.column
                    }
                }
            }],
            "properties": {
                "confidence": self.confidence,
                "standard": self.standard,
                "isNew": self.is_new,
                "fingerprint": self.fingerprint
            }
        }


class CodeReviewOrchestrator:
    """Main orchestration class"""
    
    def __init__(self, repo_dir: str, config_file: str, pr_base: str = None, pr_head: str = None):
        """
        Args:
            repo_dir: Path to iceoryx repository
            config_file: Path to review-config.yaml
            pr_base: Base branch for comparison (e.g., 'main')
            pr_head: Head branch of PR
        """
        self.repo_dir = Path(repo_dir)
        self.config_file = Path(config_file)
        self.pr_base = pr_base or "main"
        self.pr_head = pr_head or "HEAD"
        self.build_dir = self.repo_dir / "build"
        self.results_dir = self.repo_dir / ".review-results"
        
        # Load configuration
        with open(self.config_file, 'r') as f:
            self.config = yaml.safe_load(f)
        
        self.findings: List[Finding] = []
        self.changed_files: Set[str] = set()
        
    def log(self, msg: str, level: str = "INFO"):
        """Logging helper"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{timestamp}] [{level}] {msg}")
    
    def run(self) -> Dict[str, Any]:
        """Execute the full review pipeline"""
        try:
            self.log("Starting code review pipeline...")
            
            # Step 1: Prepare repository
            self._get_changed_files()
            self.log(f"Found {len(self.changed_files)} changed files")
            
            # Step 2: Build project
            self.log("Building Eclipse iceoryx with compile_commands.json...")
            self._build_project()
            
            # Step 3: Run analyzers
            self.log("Running static analyzers...")
            self._run_analyzers()
            
            # Step 4: Normalize and filter results
            self.log(f"Collected {len(self.findings)} findings")
            self._filter_new_findings()
            self.log(f"After filtering: {len(self.findings)} new findings")
            
            # Step 5: Generate output
            sarif_output = self._generate_sarif()
            summary = self._generate_summary()
            
            self.log("Code review pipeline completed successfully")
            return {
                "status": "success",
                "sarif": sarif_output,
                "summary": summary,
                "findings": [asdict(f) for f in self.findings]
            }
            
        except Exception as e:
            self.log(f"Error in review pipeline: {str(e)}", level="ERROR")
            return {
                "status": "error",
                "error": str(e)
            }
    
    def _get_changed_files(self):
        """Get list of changed files in PR"""
        if shutil.which("git") is None:
            raise RuntimeError(
                "Git command not found. Install Git and ensure it is available on PATH "
                "before running the review bot."
            )

        try:
            diff_cmd = ["git", "diff", "--name-only", f"{self.pr_base}...{self.pr_head}"]
            self.log(f"Getting changed files with: {' '.join(diff_cmd)}")
            result = subprocess.run(
                diff_cmd,
                cwd=self.repo_dir,
                capture_output=True,
                text=True,
                check=True
            )
            self.changed_files = set(result.stdout.strip().split("\n"))
            self.changed_files.discard("")  # Remove empty strings

            if not self.changed_files:
                self.log("No committed diff found; checking working tree changes...")
                result = subprocess.run(
                    ["git", "diff", "--name-only", self.pr_base],
                    cwd=self.repo_dir,
                    capture_output=True,
                    text=True,
                    check=True
                )
                self.changed_files = set(result.stdout.strip().split("\n"))
                self.changed_files.discard("")

            if not self.changed_files:
                self.log("Still no changed files; checking staged changes and untracked files")
                result = subprocess.run(
                    ["git", "status", "--porcelain"],
                    cwd=self.repo_dir,
                    capture_output=True,
                    text=True,
                    check=True
                )
                files = []
                for line in result.stdout.splitlines():
                    if not line:
                        continue
                    path = line[3:]
                    files.append(path)
                self.changed_files = set(files)

        except subprocess.CalledProcessError as e:
            self.log(f"Failed to get changed files: {e.stderr}", level="ERROR")
            raise
    
    def _build_project(self):
        """Build Eclipse iceoryx with compile_commands.json"""
        # Try Bazel first (iceoryx uses Bazel as primary build system)
        if (self.repo_dir / "WORKSPACE.bazel").exists() and shutil.which("bazel") is not None:
            self._build_with_bazel()
        # Fall back to CMake
        elif (self.repo_dir / "iceoryx_meta" / "CMakeLists.txt").exists() or (self.repo_dir / "CMakeLists.txt").exists():
            self._build_with_cmake()
        else:
            raise RuntimeError(
                f"No build system found in {self.repo_dir}. "
                "Neither WORKSPACE.bazel (Bazel) nor CMakeLists.txt (CMake) detected. "
                "Ensure you have cloned a complete iceoryx repository."
            )
    
    def _build_with_bazel(self):
        """Build using Bazel"""
        self.log("Detected Bazel build system; building with Bazel...")
        try:
            # Build and generate compile_commands.json
            bazel_cmd = [
                "bazel", "build",
                "--build_tag_filters=",
                "--test_tag_filters=",
                "//:iceoryx"
            ]
            
            self.log(f"Building with Bazel: {' '.join(bazel_cmd)}")
            subprocess.run(bazel_cmd, check=True, cwd=self.repo_dir, timeout=600)
            
            # Try to extract compile_commands from Bazel output
            self.log("Note: Bazel compilation complete; compile_commands.json may need manual extraction")
            self.log("Proceeding with source-only analysis...")
            
        except subprocess.TimeoutExpired:
            self.log("Bazel build timed out; trying CMake fallback", level="WARNING")
            if shutil.which("cmake") is not None:
                self._build_with_cmake()
        except subprocess.CalledProcessError as e:
            self.log(f"Bazel build failed: {e}; trying CMake fallback", level="WARNING")
            if shutil.which("cmake") is not None:
                self._build_with_cmake()
            else:
                self.log("CMake not available; proceeding with source-only analysis", level="WARNING")
    
    def _build_with_cmake(self):
        """Build using CMake"""
        self.log("Detected CMake build system; building with CMake...")
        if shutil.which("cmake") is None:
            raise RuntimeError(
                "CMake is not installed or not available on PATH. Install CMake or Bazel before running the review bot."
            )

        try:
            # Create build directory
            self.build_dir.mkdir(exist_ok=True)
            
            # Try root first, then iceoryx_meta
            source_dir = str(self.repo_dir)
            if not (self.repo_dir / "CMakeLists.txt").exists() and (self.repo_dir / "iceoryx_meta" / "CMakeLists.txt").exists():
                source_dir = str(self.repo_dir / "iceoryx_meta")
                self.log(f"CMakeLists.txt found in iceoryx_meta/; using that as source")
            
            # Configure with CMake
            cmake_cmd = [
                "cmake",
                "-B", str(self.build_dir),
                "-DCMAKE_EXPORT_COMPILE_COMMANDS=ON",
                "-DCMAKE_BUILD_TYPE=Release",
                source_dir
            ]
            
            self.log(f"Configuring: {' '.join(cmake_cmd)}")
            subprocess.run(cmake_cmd, check=True, cwd=self.repo_dir, timeout=300)
            
            compile_db = self.build_dir / "compile_commands.json"
            if compile_db.exists():
                self.log(f"Compilation database generated at {compile_db}")
            else:
                self.log("CMake configuration completed, but compile_commands.json was not found", level="WARNING")
            
        except subprocess.TimeoutExpired as e:
            self.log(f"CMake config timed out: {e}", level="WARNING")
        except subprocess.CalledProcessError as e:
            self.log(f"CMake configuration failed: {e}", level="ERROR")
            raise
    
    def _run_analyzers(self):
        """Run configured static analyzers"""
        
        # Collect findings from each enabled analyzer
        if self.config['analyzers']['compiler']['enabled']:
            self._run_compiler_analysis()
        
        if self.config['analyzers']['clang-tidy']['enabled']:
            self._run_clang_tidy()
        
        if self.config['analyzers']['clang-analyzer']['enabled']:
            self._run_clang_analyzer()
    
    def _run_compiler_analysis(self):
        """Extract compiler warnings from compile_commands.json"""
        self.log("Analyzing compiler warnings...")
        
        compile_db = self.build_dir / "compile_commands.json"
        if not compile_db.exists():
            self.log(f"compile_commands.json not found at {compile_db}", level="WARNING")
            return
        
        try:
            # Re-compile with verbose output to capture warnings
            result = subprocess.run(
                ["cmake", "--build", str(self.build_dir), "--verbose"],
                capture_output=True,
                text=True,
                timeout=300
            )
            
            # Parse compiler output for warnings
            self._parse_compiler_output(result.stderr + result.stdout)
            
        except subprocess.TimeoutExpired:
            self.log("Compiler analysis timed out", level="WARNING")
        except Exception as e:
            self.log(f"Compiler analysis failed: {e}", level="WARNING")
    
    def _parse_compiler_output(self, output: str):
        """Parse compiler warning messages"""
        lines = output.split("\n")
        for line in lines:
            # Simple heuristic: look for file:line:col: warning
            if "warning:" in line and ".cpp" in line or ".h" in line:
                parts = line.split(":")
                if len(parts) >= 4:
                    try:
                        file = parts[0]
                        line_no = int(parts[1])
                        col = int(parts[2])
                        message = ":".join(parts[3:]).strip()
                        
                        # Skip if not in changed files
                        if not self._is_changed_file(file):
                            continue
                        
                        finding = Finding(
                            rule_id="COMPILER-WARNING",
                            tool="compiler",
                            severity="major",
                            confidence="high",
                            file=file,
                            line=line_no,
                            column=col,
                            message=message,
                            standard="project-policy",
                            is_new=True,
                            fingerprint=f"{file}:{line_no}:{col}:{message}"
                        )
                        self.findings.append(finding)
                    except (ValueError, IndexError):
                        continue
    
    def _run_clang_tidy(self):
        """Run clang-tidy on changed files"""
        self.log("Running clang-tidy...")

        if shutil.which("clang-tidy") is None:
            self.log("clang-tidy executable not found. Install clang and clang-tidy to enable this analyzer.", level="WARNING")
            return

        compile_db = self.build_dir / "compile_commands.json"
        if not compile_db.exists():
            self.log(f"compile_commands.json not found", level="WARNING")
            self.log("Skipping clang-tidy because the compilation database is missing.", level="WARNING")
            return

        try:
            # Filter compile_commands.json to only changed files
            changed_sources = self._filter_compile_commands(compile_db)
            
            result = subprocess.run(
                [
                    "clang-tidy",
                    "-p", str(self.build_dir),
                    "--export-fixes", str(self.results_dir / "clang-tidy.yaml"),
                ] + changed_sources,
                capture_output=True,
                text=True,
                timeout=600
            )
            
            # Parse clang-tidy output
            self._parse_clang_tidy_output(result.stdout + result.stderr)
            
        except subprocess.TimeoutExpired:
            self.log("clang-tidy analysis timed out", level="WARNING")
        except Exception as e:
            self.log(f"clang-tidy analysis failed: {e}", level="WARNING")
    
    def _parse_clang_tidy_output(self, output: str):
        """Parse clang-tidy JSON output"""
        lines = output.split("\n")
        for line in lines:
            # clang-tidy output format: file:line:col: severity: message [check-name]
            if ".cpp" in line or ".h" in line:
                if "warning:" in line or "error:" in line:
                    parts = line.split(":")
                    if len(parts) >= 5:
                        try:
                            file = parts[0]
                            line_no = int(parts[1])
                            col = int(parts[2])
                            level_msg = ":".join(parts[3:]).strip()
                            
                            # Extract check name from brackets
                            import re
                            match = re.search(r'\[([^\]]+)\]', level_msg)
                            check_name = match.group(1) if match else "unknown"
                            
                            severity = "major" if "warning" in level_msg else "critical"
                            
                            finding = Finding(
                                rule_id=check_name,
                                tool="clang-tidy",
                                severity=severity,
                                confidence="high",
                                file=file,
                                line=line_no,
                                column=col,
                                message=level_msg.replace(check_name, "").strip(),
                                standard="AUTOSAR",
                                is_new=True,
                                fingerprint=f"{file}:{line_no}:{col}:{check_name}"
                            )
                            self.findings.append(finding)
                        except (ValueError, IndexError):
                            continue
    
    def _run_clang_analyzer(self):
        """Run Clang Static Analyzer"""
        self.log("Running Clang Static Analyzer...")

        if shutil.which("scan-build") is None:
            self.log("scan-build executable not found. Install clang static analyzer tools to enable this analyzer.", level="WARNING")
            return
        
        try:
            result = subprocess.run(
                [
                    "scan-build",
                    "-o", str(self.results_dir / "scan-build"),
                    "cmake", "--build", str(self.build_dir)
                ],
                capture_output=True,
                text=True,
                timeout=600,
                cwd=self.repo_dir
            )
            
            # parse_analyzer_reports would go here
            self.log("Clang analyzer results saved")
            
        except subprocess.TimeoutExpired:
            self.log("Clang analyzer timed out", level="WARNING")
        except Exception as e:
            self.log(f"Clang analyzer failed: {e}", level="WARNING")
    
    def _filter_compile_commands(self, compile_db_path: Path) -> List[str]:
        """Filter compile_commands.json to only changed files"""
        try:
            with open(compile_db_path, 'r') as f:
                compile_db = json.load(f)
            
            changed = []
            for entry in compile_db:
                file_path = entry.get("file", "")
                if any(file_path.endswith(cf) for cf in self.changed_files):
                    changed.append(file_path)
            
            return changed[:20]  # Limit to first 20 to avoid timeout
            
        except Exception as e:
            self.log(f"Error filtering compile_commands: {e}", level="WARNING")
            return []
    
    def _is_changed_file(self, file_path: str) -> bool:
        """Check if file is in changed files"""
        return any(file_path.endswith(cf) for cf in self.changed_files)
    
    def _filter_new_findings(self):
        """Keep only findings from changed lines"""
        if not self.config['review_policy'].get('only_new_findings', True):
            return
        
        # Deduplicate by fingerprint
        seen = {}
        unique_findings = []
        for finding in self.findings:
            if finding.fingerprint not in seen:
                unique_findings.append(finding)
                seen[finding.fingerprint] = True
        
        self.findings = unique_findings
    
    def _generate_sarif(self) -> Dict[str, Any]:
        """Generate SARIF report"""
        sarif = {
            "version": "2.1.0",
            "$schema": "https://raw.githubusercontent.com/oasis-tcs/sarif-spec/master/Schemata/sarif-schema-2.1.0.json",
            "runs": [
                {
                    "tool": {
                        "driver": {
                            "name": "iceoryx-code-review-agent",
                            "version": "1.0.0",
                            "informationUri": "https://github.com/eclipse-iceoryx/iceoryx"
                        }
                    },
                    "results": [f.to_sarif_result() for f in self.findings]
                }
            ]
        }
        
        # Save SARIF file
        self.results_dir.mkdir(exist_ok=True)
        sarif_path = self.results_dir / "review-results.sarif"
        with open(sarif_path, 'w') as f:
            json.dump(sarif, f, indent=2)
        
        self.log(f"SARIF report saved to {sarif_path}")
        return sarif

    def _generate_excel(self, output_path: Path):
        """Generate an Excel report from findings"""
        if Workbook is None or get_column_letter is None:
            raise RuntimeError(
                "openpyxl is required to generate Excel reports. Install it with: pip install openpyxl"
            )

        self.results_dir.mkdir(exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Code Review Findings"

        headers = [
            "File",
            "Line",
            "Column",
            "Severity",
            "Rule ID",
            "Tool",
            "Confidence",
            "Standard",
            "Message",
            "Is New",
            "Fingerprint"
        ]
        ws.append(headers)

        for finding in sorted(self.findings, key=lambda f: (f.file, f.line, f.column)):
            ws.append([
                finding.file,
                finding.line,
                finding.column,
                finding.severity,
                finding.rule_id,
                finding.tool,
                finding.confidence,
                finding.standard,
                finding.message,
                finding.is_new,
                finding.fingerprint
            ])

        for idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 18

        wb.save(output_path)
        self.log(f"Excel report saved to {output_path}")
    
    def _generate_summary(self) -> Dict[str, Any]:
        """Generate review summary"""
        by_severity = {}
        for f in self.findings:
            if f.severity not in by_severity:
                by_severity[f.severity] = []
            by_severity[f.severity].append(f)
        
        return {
            "total_findings": len(self.findings),
            "by_severity": {
                sev: len(findings) for sev, findings in by_severity.items()
            },
            "changed_files": len(self.changed_files),
            "timestamp": datetime.now().isoformat()
        }


def main():
    parser = argparse.ArgumentParser(description="Code Review Orchestrator for iceoryx")
    parser.add_argument("--repo", required=True, help="Path to iceoryx repository")
    parser.add_argument("--config", required=True, help="Path to review-config.yaml")
    parser.add_argument("--pr-base", default="main", help="Base branch for PR comparison")
    parser.add_argument("--pr-head", default="HEAD", help="Head branch for PR comparison")
    parser.add_argument("--output", default="review-results.json", help="Output file for results")
    parser.add_argument("--xlsx-output", default=None, help="Optional Excel output file (.xlsx)")
    
    args = parser.parse_args()
    
    orchestrator = CodeReviewOrchestrator(
        repo_dir=args.repo,
        config_file=args.config,
        pr_base=args.pr_base,
        pr_head=args.pr_head
    )
    
    results = orchestrator.run()
    
    # Save results
    with open(args.output, 'w') as f:
        json.dump(results, f, indent=2)

    if args.xlsx_output:
        try:
            orchestrator._generate_excel(Path(args.xlsx_output))
        except Exception as e:
            orchestrator.log(f"Failed to generate Excel report: {e}", level="ERROR")
            results = {
                "status": "error",
                "error": str(e)
            }
            with open(args.output, 'w') as f:
                json.dump(results, f, indent=2)

    print(f"\nResults saved to {args.output}")
    if args.xlsx_output:
        print(f"Excel report saved to {args.xlsx_output}")
    
    # Exit with appropriate code
    sys.exit(0 if results['status'] == 'success' else 1)


if __name__ == "__main__":
    main()
