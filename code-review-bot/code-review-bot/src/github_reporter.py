#!/usr/bin/env python3
"""
GitHub Integration for Code Review Bot
Posts findings as PR comments and checks
"""

import json
import os
import sys
from pathlib import Path
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
from datetime import datetime

try:
    from github import Github, GithubException
except ImportError:
    print("ERROR: PyGithub not installed. Install with: pip install PyGithub")
    sys.exit(1)

try:
    import requests
except Exception:
    requests = None


@dataclass
class ReviewComment:
    """A single review comment on a PR"""
    file: str
    line: int
    message: str
    side: str = "RIGHT"  # RIGHT for new changes, LEFT for base


class GitHubReportGenerator:
    """Generates GitHub PR reports from review findings"""
    
    def __init__(self, github_token: Optional[str] = None, repo_name: Optional[str] = None):
        """
        Args:
            github_token: GitHub token (or use GITHUB_TOKEN env var)
            repo_name: Repository name as "owner/repo" (or use GITHUB_REPOSITORY env var)
        """
        self.github_token = github_token or os.getenv("GITHUB_TOKEN")
        self.repo_name = repo_name or os.getenv("GITHUB_REPOSITORY")
        self.pr_number = int(os.getenv("GITHUB_PR_NUMBER", 0))
        self.run_id = os.getenv("GITHUB_RUN_ID", "unknown")

        if self.github_token and self.repo_name:
            self.github = Github(self.github_token)
            self.repo = self.github.get_repo(self.repo_name)
            self.pull = None
            if self.pr_number > 0:
                self.pull = self.repo.get_pull(self.pr_number)
        else:
            self.github = None
            self.repo = None
            self.pull = None
        
        self.review_comments: List[ReviewComment] = []
        self.summary_checks: Dict[str, Any] = {}
    
    def load_results(self, results_file: str):
        """Load review results from JSON file"""
        with open(results_file, 'r') as f:
            results = json.load(f)
        
        if results.get('status') != 'success':
            print(f"ERROR: Review failed with status: {results.get('status')}")
            return
        
        findings = results.get('findings', [])
        summary = results.get('summary', {})
        
        self._process_findings(findings, summary)
    
    def _process_findings(self, findings: List[Dict], summary: Dict):
        """Convert findings to GitHub comments"""
        
        # Group by severity
        by_severity = {}
        for finding in findings:
            sev = finding['severity']
            if sev not in by_severity:
                by_severity[sev] = []
            by_severity[sev].append(finding)
        
        # Create summary
        self.summary_checks = {
            'total': len(findings),
            'by_severity': {sev: len(fs) for sev, fs in by_severity.items()},
            'files_reviewed': summary.get('changed_files', 0)
        }
        
        # Limit inline comments to high-severity findings only
        for finding in findings:
            if finding['severity'] in ['critical', 'major']:
                comment = self._finding_to_comment(finding)
                self.review_comments.append(comment)
    
    def _finding_to_comment(self, finding: Dict[str, Any]) -> ReviewComment:
        """Convert a finding to a PR comment"""
        
        rule_id = finding.get('rule_id', 'UNKNOWN')
        severity = finding.get('severity', 'info').upper()
        confidence = finding.get('confidence', 'unknown').upper()
        standard = finding.get('standard', 'project-policy')
        message = finding.get('message', 'No description')
        
        comment_text = f"""**[{severity}]** {rule_id}
**Standard:** {standard}
**Confidence:** {confidence}

{message}

---
*Detected by iceoryx-code-review-agent ([Run #{self.run_id}](https://github.com/{self.repo_name}/actions/runs/{self.run_id}))*
"""
        
        return ReviewComment(
            file=finding['file'],
            line=finding['line'],
            message=comment_text,
            side="RIGHT"
        )
    
    def post_summary_check(self):
        """Post a check run with summary"""
        if not self.pull:
            print("No PR context; skipping check run")
            return
        
        status = "completed"
        conclusion = "success"
        
        # Determine conclusion based on findings
        critical_count = self.summary_checks.get('by_severity', {}).get('critical', 0)
        major_count = self.summary_checks.get('by_severity', {}).get('major', 0)
        
        if critical_count > 0:
            conclusion = "failure"
        elif major_count > 3:
            conclusion = "neutral"
        
        summary_title = f"Code Review: {self.summary_checks['total']} findings"
        summary_text = f"""
## Eclipse iceoryx Code Review Report

**Summary:**
- **Total Findings:** {self.summary_checks['total']}
- **Critical:** {self.summary_checks.get('by_severity', {}).get('critical', 0)}
- **Major:** {self.summary_checks.get('by_severity', {}).get('major', 0)}
- **Minor:** {self.summary_checks.get('by_severity', {}).get('minor', 0)}
- **Files Reviewed:** {self.summary_checks['files_reviewed']}

**Standards Checked:**
- MISRA C++ (via internal mapping)
- AUTOSAR C++14
- Project-specific policies
- Clang compiler warnings

Only high-severity issues are shown inline. See run details for complete results.
"""
        
        try:
            check = self.repo.create_check_run(
                name="iceoryx-code-review",
                head_sha=self.pull.head.sha,
                status=status,
                conclusion=conclusion,
                output={
                    "title": summary_title,
                    "summary": summary_text,
                    "annotations": self._generate_check_annotations()
                }
            )
            print(f"✓ Posted check run: {check.html_url}")
            
        except GithubException as e:
            print(f"ERROR posting check run: {e}")
    
    def _generate_check_annotations(self) -> List[Dict[str, Any]]:
        """Generate check run annotations from findings"""
        annotations = []
        
        for comment in self.review_comments[:50]:  # Limit to 50 annotations (GitHub API limit)
            annotation_level = "warning"
            if "CRITICAL" in comment.message:
                annotation_level = "failure"
            
            annotations.append({
                "path": comment.file,
                "start_line": comment.line,
                "end_line": comment.line,
                "annotation_level": annotation_level,
                "message": comment.message.split('\n')[0]  # First line only
            })
        
        return annotations
    
    def post_inline_comments(self):
        """Post inline comments on changed lines"""
        if not self.pull:
            print("No PR context; skipping inline comments")
            return
        
        try:
            # Get PR files to check line availability
            pr_files = {f.filename: f for f in self.pull.get_files()}
            
            posted = 0
            for comment in self.review_comments:
                if comment.file not in pr_files:
                    print(f"  Skipping {comment.file} (not in PR)")
                    continue
                
                try:
                    self.pull.create_review_comment(
                        body=comment.message,
                        commit_id=self.pull.head.sha,
                        path=comment.file,
                        line=comment.line,
                        side=comment.side
                    )
                    posted += 1
                    print(f"  ✓ Posted comment on {comment.file}:{comment.line}")
                    
                except GithubException as e:
                    print(f"  ✗ Failed to post comment on {comment.file}: {e}")
            
            print(f"✓ Posted {posted}/{len(self.review_comments)} inline comments")
            
        except GithubException as e:
            print(f"ERROR posting inline comments: {e}")
    
    def post_pr_comment(self, body: str):
        """Post a general comment on the PR"""
        if not self.pull:
            print("No PR context; skipping PR comment")
            return
        
        try:
            self.pull.create_issue_comment(body)
            print("✓ Posted PR comment")
        except GithubException as e:
            print(f"ERROR posting PR comment: {e}")
    
    def generate_report(self, results_file: str, output_md: str = "review-report.md"):
        """Generate a standalone markdown report"""
        findings, summary = self._load_results(results_file)
        self._generate_markdown_report(findings, summary, output_md)

    def generate_xlsx(self, results_file: str, output_xlsx: str = "review-report.xlsx"):
        """Generate a standalone Excel report"""
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("ERROR: openpyxl is not installed. Install with: pip install openpyxl")
            return

        findings, summary = self._load_results(results_file)

        wb = Workbook()
        ws = wb.active
        ws.title = "Code Review Findings"

        headers = [
            "File",
            "Line",
            "Column",
            "Severity",
            "Rule ID",
            "Confidence",
            "Standard",
            "Message",
            "Is New",
            "Fingerprint"
        ]
        ws.append(headers)

        for finding in findings:
            ws.append([
                finding.get('file', ''),
                finding.get('line', ''),
                finding.get('column', ''),
                finding.get('severity', ''),
                finding.get('rule_id', ''),
                finding.get('confidence', ''),
                finding.get('standard', ''),
                finding.get('message', ''),
                finding.get('is_new', ''),
                finding.get('fingerprint', '')
            ])

        for col_idx, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        wb.save(output_xlsx)
        print(f"✓ Generated Excel report: {output_xlsx}")

    def _load_results(self, results_file: str):
        with open(results_file, 'r') as f:
            results = json.load(f)
        findings = results.get('findings', [])
        summary = results.get('summary', {})
        return findings, summary

    def _generate_markdown_report(self, findings: List[Dict], summary: Dict, output_md: str):
        by_file = {}
        for finding in findings:
            f = finding['file']
            if f not in by_file:
                by_file[f] = []
            by_file[f].append(finding)

        md_content = f"""# Code Review Report
Generated: {datetime.now().isoformat()}

## Summary
- **Total Findings:** {len(findings)}
- **Files Reviewed:** {summary.get('changed_files', 0)}
- **Critical:** {summary.get('by_severity', {}).get('critical', 0)}
- **Major:** {summary.get('by_severity', {}).get('major', 0)}
- **Minor:** {summary.get('by_severity', {}).get('minor', 0)}

## Findings by File

"""

        for file_name in sorted(by_file.keys()):
            file_findings = by_file[file_name]
            md_content += f"\n### {file_name}\n\n"

            for finding in sorted(file_findings, key=lambda x: x['line']):
                md_content += f"""**Line {finding['line']} [{finding['severity'].upper()}]** - {finding['rule_id']}
- Standard: {finding.get('standard', 'N/A')}
- Message: {finding['message']}

"""

        with open(output_md, 'w') as f:
            f.write(md_content)

        print(f"✓ Generated report: {output_md}")


def main():
    import argparse

    parser = argparse.ArgumentParser(description="GitHub Integration for Code Review Bot")
    parser.add_argument("--results", required=True, help="Path to review-results.json")
    parser.add_argument("--post-check", action="store_true", help="Post check run to PR")
    parser.add_argument("--post-comments", action="store_true", help="Post inline comments to PR")
    parser.add_argument("--generate-report", action="store_true", help="Generate markdown report")
    parser.add_argument("--output", default="review-report.md", help="Report output file")

    args = parser.parse_args()

    try:
        reporter = None
        if os.getenv("GITHUB_ACTIONS") == "true":
            reporter = GitHubReportGenerator()
            reporter.load_results(args.results)

            if args.post_check:
                print("Posting check run...")
                reporter.post_summary_check()

            if args.post_comments:
                print("Posting inline comments...")
                reporter.post_inline_comments()
        else:
            print("Not running in GitHub Actions context; report mode only")

        if args.generate_report and reporter is not None:
            print(f"Generating markdown report to {args.output}...")
            reporter.generate_report(args.results, args.output)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
